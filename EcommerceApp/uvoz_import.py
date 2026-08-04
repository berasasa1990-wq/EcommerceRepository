"""
Uvoz artikala iz Excel-a (format kao „Uvoz dd.mm.yyyy.xlsx”).

Očekivane kolone u redu zaglavlja artikala:
  Artikal | Kolicina | Fakturna | Nabavna | Vpc netto | Mpc brutto | …

Pravila:
  - Kolicina > 0 i Artikal postoji u bazi (naziv 100% isti) → na stanju + cijena Mpc brutto
  - Artikal ne postoji → kreiraj, na stanju, cijena Mpc brutto
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

from .models import Product, Uvoz, UvozStavka

# Kolone u Excelu (header red) — mapiranje po nazivu
HEADER_ALIASES = {
    'artikal': 'artikal',
    'kolicina': 'kolicina',
    'količina': 'kolicina',
    'fakturna': 'fakturna',
    'nabavna': 'nabavna',
    'vpc netto': 'vpc_netto',
    'vpc netto ': 'vpc_netto',
    'mpc brutto': 'mpc_brutto',
    'mpc brutto ': 'mpc_brutto',
    'vpc marza': 'vpc_marza',
    'vpc marža': 'vpc_marza',
    'ukupno fakturna': 'ukupno_fakturna',
}


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _normalize_header(value) -> str:
    return re.sub(r'\s+', ' ', _cell_str(value).casefold()).strip()


def parse_money(value) -> Decimal | None:
    """
    '  5.00 KM ' / '5,50' / 5.0 → Decimal
    """
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            return None
    text = _cell_str(value)
    if not text:
        return None
    text = text.replace('KM', '').replace('km', '').replace('\xa0', ' ').strip()
    text = text.replace(' ', '')
    if ',' in text and '.' in text:
        if text.rfind(',') > text.rfind('.'):
            text = text.replace('.', '').replace(',', '.')
        else:
            text = text.replace(',', '')
    else:
        text = text.replace(',', '.')
    text = re.sub(r'[^0-9.\-]', '', text)
    if not text or text in ('.', '-', '-.'):
        return None
    try:
        return Decimal(text).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return None


def parse_qty(value) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value))
    text = _cell_str(value).replace(',', '.')
    text = re.sub(r'[^0-9.\-]', '', text)
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows):
        if not row:
            continue
        mapping = {}
        for col_i, cell in enumerate(row):
            key = HEADER_ALIASES.get(_normalize_header(cell))
            if key and key not in mapping:
                mapping[key] = col_i
        if 'artikal' in mapping and 'mpc_brutto' in mapping:
            return idx, mapping
    raise ValueError(
        'Nije pronađen red zaglavlja sa kolonama „Artikal” i „Mpc brutto”. '
        'Provjeri da Excel ima isti format kao uvoz (Artikal | Kolicina | … | Mpc brutto).',
    )


def read_uvoz_rows_from_workbook(wb) -> list[dict]:
    ws = wb.active
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    header_idx, colmap = _find_header_row(rows)

    items = []
    for row in rows[header_idx + 1:]:
        if not row:
            continue
        name = _cell_str(row[colmap['artikal']] if colmap['artikal'] < len(row) else None)
        if not name:
            continue
        if _normalize_header(name) == 'artikal':
            continue

        def _col(key):
            i = colmap.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        qty = parse_qty(_col('kolicina'))
        price = parse_money(_col('mpc_brutto'))
        marza_raw = _col('vpc_marza')
        marza = None
        if marza_raw is not None and marza_raw != '':
            try:
                marza = Decimal(str(marza_raw))
            except (InvalidOperation, ValueError):
                marza = parse_money(marza_raw)

        items.append({
            'artikal': name[:200],
            'kolicina': qty,
            'mpc_brutto': price,
            'fakturna': parse_money(_col('fakturna')),
            'nabavna': parse_money(_col('nabavna')),
            'vpc_netto': parse_money(_col('vpc_netto')),
            'vpc_marza': marza,
            'ukupno_fakturna': parse_money(_col('ukupno_fakturna')),
        })
    return items


def parse_uvoz_excel(file_obj) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            'openpyxl nije instaliran na serveru. '
            'Dodaj paket u requirements i redeploy: pip install "openpyxl>=3.1.0"',
        ) from exc

    name = (getattr(file_obj, 'name', '') or '').lower()
    if name and not (name.endswith('.xlsx') or name.endswith('.xlsm')):
        raise ValueError('Podržan je Excel .xlsx (npr. Uvoz 16.04.2026.xlsx).')

    raw = file_obj.read()
    if hasattr(file_obj, 'seek'):
        try:
            file_obj.seek(0)
        except Exception:
            pass
    if not raw:
        raise ValueError('Prazan fajl.')

    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    try:
        return read_uvoz_rows_from_workbook(wb)
    finally:
        wb.close()


def _find_product_exact_name(name: str) -> Product | None:
    """100% isti naziv (nakon strip)."""
    name = (name or '').strip()
    if not name:
        return None
    product = Product.objects.filter(naziv=name).first()
    if product:
        return product
    collapsed = re.sub(r'\s+', ' ', name)
    if collapsed != name:
        product = Product.objects.filter(naziv=collapsed).first()
        if product:
            return product
    for p in Product.objects.filter(naziv__iexact=name).only('id', 'naziv')[:5]:
        if (p.naziv or '').strip() == name:
            return p
    return None


def _qty_int(qty) -> int:
    if qty is None:
        return 0
    try:
        if hasattr(qty, 'to_integral_value') and qty == qty.to_integral_value():
            n = int(qty)
        else:
            n = int(qty)
    except Exception:
        n = 0
    return max(n, 0)


def apply_row_to_product(row: dict, *, apply_stock_rules: bool = True) -> tuple[str, Product | None, str]:
    """
    Primijeni jedan red na Product.
    Vraća (status, product|None, poruka)
    status: created | updated | skipped | error
    """
    name = (row.get('artikal') or row.get('artikal_naziv') or '').strip()
    qty = row.get('kolicina')
    price = row.get('mpc_brutto')

    if not name:
        return UvozStavka.Status.SKIPPED, None, 'Prazan naziv'

    if apply_stock_rules:
        if qty is None or qty <= 0:
            return UvozStavka.Status.SKIPPED, None, 'Količina ≤ 0'
        if price is None or price <= 0:
            return UvozStavka.Status.SKIPPED, None, 'Nema Mpc brutto'

    qty_int = _qty_int(qty) or 1
    if price is None or price <= 0:
        return UvozStavka.Status.SKIPPED, None, 'Nema Mpc brutto'

    try:
        with transaction.atomic():
            product = _find_product_exact_name(name)
            if product:
                fields = []
                old_price = product.cijena
                if product.cijena != price:
                    product.cijena = price
                    if product.akcija_postotak:
                        product.akcijska_cijena = None
                        fields.append('akcijska_cijena')
                    fields.append('cijena')
                if not product.na_stanju:
                    product.na_stanju = True
                    fields.append('na_stanju')
                if product.stanje != qty_int:
                    product.stanje = qty_int
                    fields.append('stanje')
                if not product.aktivan:
                    product.aktivan = True
                    fields.append('aktivan')
                if fields:
                    product.save()
                    return (
                        UvozStavka.Status.UPDATED,
                        product,
                        f'Ažurirano: cijena {old_price}→{price} KM, kol. {qty_int}',
                    )
                return UvozStavka.Status.UPDATED, product, 'Već usklađeno'
            product = Product(
                naziv=name[:200],
                cijena=price,
                na_stanju=True,
                stanje=qty_int,
                aktivan=True,
                prikazi_na_pocetnoj=True,
            )
            if not product.slug:
                product.slug = slugify(name)[:180] or 'artikal'
            product.save()
            return (
                UvozStavka.Status.CREATED,
                product,
                f'Kreirano: {price} KM, kol. {qty_int}',
            )
    except Exception as exc:
        return UvozStavka.Status.ERROR, None, str(exc)


def apply_uvoz_import(rows: list[dict], *, dry_run: bool = False) -> dict:
    """Primijeni uvoz (bez snimanja Uvoz modela). Samo redovi s kolicina > 0."""
    stats = {
        'rows_total': len(rows),
        'rows_qty_positive': 0,
        'updated': 0,
        'created': 0,
        'skipped_qty': 0,
        'skipped_no_price': 0,
        'errors': [],
        'details': [],
    }

    def add_detail(msg: str):
        if len(stats['details']) < 100:
            stats['details'].append(msg)

    for row in rows:
        name = (row.get('artikal') or '').strip()
        qty = row.get('kolicina')
        price = row.get('mpc_brutto')
        if not name:
            continue
        if qty is None or qty <= 0:
            stats['skipped_qty'] += 1
            continue
        stats['rows_qty_positive'] += 1
        if price is None or price <= 0:
            stats['skipped_no_price'] += 1
            add_detail(f'Preskočeno (nema Mpc brutto): {name}')
            continue

        if dry_run:
            existing = _find_product_exact_name(name)
            if existing:
                stats['updated'] += 1
                add_detail(f'Ažuriranje (dry-run): {name} → {price} KM')
            else:
                stats['created'] += 1
                add_detail(f'Novi (dry-run): {name} → {price} KM')
            continue

        status, _product, msg = apply_row_to_product(row)
        if status == UvozStavka.Status.CREATED:
            stats['created'] += 1
        elif status == UvozStavka.Status.UPDATED:
            stats['updated'] += 1
        elif status == UvozStavka.Status.ERROR:
            stats['errors'].append(f'{name}: {msg}')
        add_detail(f'{msg}: {name}' if status != UvozStavka.Status.ERROR else f'Greška: {name} — {msg}')

    return stats


def create_uvoz_from_rows(
    rows: list[dict],
    *,
    fajl_naziv: str = '',
    naziv: str = '',
    user=None,
    apply_to_products: bool = True,
) -> tuple[Uvoz, dict]:
    """
    Snimi Uvoz + stavke, po želji primijeni na artikle.
    """
    if not naziv:
        base = (fajl_naziv or '').rsplit('.', 1)[0].strip()
        naziv = base or f'Uvoz {timezone.localtime().strftime("%d.%m.%Y. %H:%M")}'

    stats = {
        'rows_total': len(rows),
        'rows_qty_positive': 0,
        'updated': 0,
        'created': 0,
        'skipped_qty': 0,
        'skipped_no_price': 0,
        'errors': [],
        'details': [],
    }

    with transaction.atomic():
        uvoz = Uvoz.objects.create(
            naziv=naziv[:200],
            fajl_naziv=(fajl_naziv or '')[:255],
            kreirao=user if getattr(user, 'is_authenticated', False) else None,
            broj_redova=len(rows),
        )
        stavke = []
        for i, row in enumerate(rows):
            name = (row.get('artikal') or '').strip()[:200]
            if not name:
                continue
            stavka = UvozStavka(
                uvoz=uvoz,
                artikal_naziv=name,
                kolicina=row.get('kolicina'),
                fakturna=row.get('fakturna'),
                nabavna=row.get('nabavna'),
                vpc_netto=row.get('vpc_netto'),
                mpc_brutto=row.get('mpc_brutto'),
                vpc_marza=row.get('vpc_marza'),
                ukupno_fakturna=row.get('ukupno_fakturna'),
                redoslijed=i,
                status=UvozStavka.Status.PENDING,
            )
            stavke.append(stavka)
        UvozStavka.objects.bulk_create(stavke)

        # reload with pks
        for stavka in uvoz.stavke.all():
            row = {
                'artikal': stavka.artikal_naziv,
                'kolicina': stavka.kolicina,
                'mpc_brutto': stavka.mpc_brutto,
            }
            qty = stavka.kolicina
            price = stavka.mpc_brutto

            if qty is None or qty <= 0:
                stats['skipped_qty'] += 1
                stavka.status = UvozStavka.Status.SKIPPED
                stavka.poruka = 'Količina ≤ 0'
                stavka.save(update_fields=['status', 'poruka'])
                continue

            stats['rows_qty_positive'] += 1

            if not apply_to_products:
                stavka.status = UvozStavka.Status.PENDING
                stavka.poruka = 'Sačuvano — nije primijenjeno na sajt'
                stavka.save(update_fields=['status', 'poruka'])
                continue

            if price is None or price <= 0:
                stats['skipped_no_price'] += 1
                stavka.status = UvozStavka.Status.SKIPPED
                stavka.poruka = 'Nema Mpc brutto'
                stavka.save(update_fields=['status', 'poruka'])
                continue

            status, product, msg = apply_row_to_product(row)
            stavka.status = status
            stavka.poruka = (msg or '')[:300]
            if product:
                stavka.product = product
            stavka.save(update_fields=['status', 'poruka', 'product'])

            if status == UvozStavka.Status.CREATED:
                stats['created'] += 1
            elif status == UvozStavka.Status.UPDATED:
                stats['updated'] += 1
            elif status == UvozStavka.Status.ERROR:
                stats['errors'].append(f'{stavka.artikal_naziv}: {msg}')

            if len(stats['details']) < 100:
                stats['details'].append(f'{stavka.artikal_naziv}: {msg}')

        uvoz.broj_azurirano = stats['updated']
        uvoz.broj_kreirano = stats['created']
        uvoz.broj_preskoceno = stats['skipped_qty'] + stats['skipped_no_price']
        uvoz.broj_redova = uvoz.stavke.count()
        uvoz.log_detalji = stats['details']
        uvoz.save()

    stats['uvoz_id'] = uvoz.pk
    stats['uvoz'] = uvoz
    return uvoz, stats


def reapply_stavka(stavka: UvozStavka) -> UvozStavka:
    """Ponovo primijeni stavku na Product (npr. nakon ručne izmjene)."""
    row = {
        'artikal': stavka.artikal_naziv,
        'kolicina': stavka.kolicina,
        'mpc_brutto': stavka.mpc_brutto,
    }
    status, product, msg = apply_row_to_product(row, apply_stock_rules=True)
    stavka.status = status
    stavka.poruka = (msg or '')[:300]
    if product:
        stavka.product = product
    stavka.save()
    return stavka
